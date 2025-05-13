#!/usr/bin/env python3
# Shiny GUI application for HFE mutation analysis

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
import re
from datetime import datetime
import io
import base64
from scipy.stats import chi2_contingency

# Import Shiny for Python
from shiny import App, ui, render, reactive
import shinyswatch

# Define reusable functions from the existing scripts
def clean_dataset(df, remove_blank_validovany=True, remove_blank_diagnoza=True, remove_blank_hfe=True, remove_second_column=True):
    """
    Clean the dataset by fixing ID format and removing rows with blank values in specified columns,
    following same approach as format_dataset.py script
    """
    results = []
    results.append("Starting dataset cleaning...")
    results.append(f"Initial rows: {len(df)}")
    
    # Find ID column (likely first column)
    if 'id' in df.columns:
        id_col = 'id'
    elif any(col.lower() == 'id' for col in df.columns):
        id_col = next(col for col in df.columns if col.lower() == 'id')
    else:
        # Assuming the first column is the ID
        id_col = df.columns[0]
    
    results.append(f"Using '{id_col}' as the ID column")
    
    # Display original column information
    results.append("Original columns:")
    for i, col in enumerate(df.columns):
        results.append(f"Column {i}: '{col}'")
    
    # Remove the second column (index 1) if it exists and removal is requested
    if remove_second_column and len(df.columns) > 1:
        second_col = df.columns[1]
        df = df.drop(columns=[second_col])
        results.append(f"Removed column at index 1: '{second_col}'")
    
    # Rename columns at indices 2 and 4 if they exist
    if len(df.columns) > 4:  # Ensure columns exist
        col_index_2 = df.columns[2]
        col_index_4 = df.columns[4]
        
        # Create a mapping for renaming
        rename_mapping = {
            col_index_2: 'cas_validacie',  # Rename column at index 2 to "cas_validacie"
            col_index_4: 'cas_prijmu'      # Rename column at index 4 to "cas_prijmu"
        }
        
        # Apply the renaming
        df = df.rename(columns=rename_mapping)
        results.append(f"Renamed columns:")
        results.append(f"- Column at index 2: '{col_index_2}' → 'cas_validacie'")
        results.append(f"- Column at index 4: '{col_index_4}' → 'cas_prijmu'")
    
    # Display updated column information
    results.append("Columns after modifications:")
    for i, col in enumerate(df.columns):
        results.append(f"Column {i}: '{col}'")
    
    # Check for problematic ID values
    total_rows = len(df)
    null_ids = df[id_col].isnull().sum()
    empty_str_ids = (df[id_col] == '').sum() if df[id_col].dtype == object else 0
    zero_ids = (df[id_col] == 0).sum() if df[id_col].dtype in [np.int64, np.float64] else 0
    
    results.append(f"ID column diagnostics:")
    results.append(f"Total rows: {total_rows}")
    results.append(f"Null/NaN IDs: {null_ids}")
    results.append(f"Empty string IDs: {empty_str_ids}")
    results.append(f"Zero IDs: {zero_ids}")
    results.append(f"ID column dtype: {df[id_col].dtype}")
    
    # Initial ID conversion for numeric columns
    if df[id_col].dtype in [np.int64, np.float64]:
        # For numeric columns, only convert non-NaN values to string
        df[id_col] = df[id_col].astype('Int64')  # nullable integer type
        df[id_col] = df[id_col].astype(str)
        # Replace 'nan' or '<NA>' strings with None
        df[id_col] = df[id_col].replace(['nan', '<NA>'], None)
        results.append("Converted ID column to string format and fixed NaN values")
    
    # Determine if IDs are in old format (YMMDDNNNN - 9 digits) or new format (YYMMDDNNNN - 10 digits)
    sample_ids = df[id_col].dropna().head(10).tolist()
    id_lengths = [len(str(id_val)) for id_val in sample_ids if str(id_val) != 'nan' and str(id_val) != '']
    
    old_format = False
    if id_lengths and max(id_lengths) <= 9:
        old_format = True
        results.append("Detected old ID format (YMMDDNNNN). Will convert to new format (YYMMDDNNNN).")
    else:
        results.append("Assuming IDs are already in the new format (YYMMDDNNNN) or no consistent format detected.")
    
    # Apply zfill to non-null values to preserve leading zeros
    mask = df[id_col].notna() & (df[id_col] != '')
    if any(mask):
        # Use zfill with the appropriate length based on format
        zfill_length = 9 if old_format else 10
        df.loc[mask, id_col] = df.loc[mask, id_col].str.zfill(zfill_length)
        results.append(f"Applied {zfill_length}-digit zero padding to IDs")
    
    # Find reception date column
    datum_prijmu_col = None
    for col in df.columns:
        if any(term in str(col).lower() for term in ['datum_prijmu', 'prijem', 'date', 'datum']):
            if 'cas' not in str(col).lower():  # Exclude time columns
                datum_prijmu_col = col
                break
    
    if datum_prijmu_col:
        results.append(f"Using '{datum_prijmu_col}' as the reception date column")
    else:
        results.append("Could not identify the reception date column")
    
    # Define function to extract date parts and construct ID prefix
    def extract_date_parts(date_val):
        if pd.isna(date_val):
            return None
        
        try:
            # If already a datetime
            if isinstance(date_val, (pd.Timestamp, datetime)):
                dt = date_val
            else:
                # Try parsing string with different formats
                for fmt in ['%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d']:
                    try:
                        dt = pd.to_datetime(date_val, format=fmt)
                        break
                    except:
                        continue
                else:
                    # If none of the specific formats worked, let pandas guess
                    dt = pd.to_datetime(date_val)
            
            # Extract parts - get last TWO digits of year
            year_last_two_digits = str(dt.year)[-2:]
            month = f"{dt.month:02d}"  # Month with leading zero
            day = f"{dt.day:02d}"  # Day with leading zero
            
            # Construct ID prefix: YYMMDD (6 characters)
            id_prefix = f"{year_last_two_digits}{month}{day}"
            return id_prefix
        except Exception as e:
            results.append(f"Error extracting date parts: {e}")
            return None
    
    # Get rows with missing IDs - include empty strings and NaN values
    missing_id_rows = df[(df[id_col].isna()) | (df[id_col] == '')]
    
    # Reconstruct missing IDs if we have reception date
    if len(missing_id_rows) > 0 and datum_prijmu_col is not None:
        results.append("RECONSTRUCTING MISSING IDs")
        
        # Function to clean and convert time to an integer (for sorting)
        def time_to_int(time_str):
            if pd.isna(time_str):
                return 999999  # High value for missing times (will be ordered last)
            # Try to clean and standardize the time string
            cleaned = str(time_str).replace(':', '').replace(' ', '').strip()
            # Extract just the digits
            digits = re.sub(r'\D', '', cleaned)
            if digits:
                return int(digits)
            return 999999
        
        # Track how many IDs were successfully reconstructed
        reconstructed_count = 0
        solo_reconstructed_count = 0
        
        # Find time column for sorting
        cas_prijmu_col = None
        for col in df.columns:
            if 'cas' in str(col).lower() and ('prijm' in str(col).lower() or 'prij' in str(col).lower()):
                cas_prijmu_col = col
                break
        
        if cas_prijmu_col:
            results.append(f"Using '{cas_prijmu_col}' as reception time column for sorting")
        
        # Process each missing ID row
        for idx, row in missing_id_rows.iterrows():
            reception_date = row[datum_prijmu_col]
            
            if pd.isna(reception_date):
                results.append(f"Row {idx}: Cannot reconstruct ID - missing reception date")
                continue
            
            # Find all rows with the same reception date
            same_date_rows = df[df[datum_prijmu_col] == reception_date].copy()
            
            # METHOD 1: If multiple records with same date, use ranking by time
            if len(same_date_rows) >= 2 and cas_prijmu_col is not None:
                # Extract valid IDs from same date (to use as template)
                valid_ids = same_date_rows[(same_date_rows[id_col].notna()) & (same_date_rows[id_col] != '')][id_col]
                
                if not valid_ids.empty:
                    # Extract a template ID - check if we need to use old or new format logic
                    template_id = valid_ids.iloc[0]
                    if old_format and len(template_id) == 9:
                        # For old format, construct new format prefix
                        id_prefix = extract_date_parts(reception_date)
                        if not id_prefix:
                            results.append(f"Row {idx}: Failed to extract date parts from {reception_date}")
                            continue
                    else:
                        # For new format, use first 6 characters (YYMMDD)
                        id_prefix = template_id[:6]  # Extract YYMMDD part
                    
                    # Add time processing for ordering
                    same_date_rows['time_int'] = same_date_rows[cas_prijmu_col].apply(time_to_int)
                    
                    # Add original row position as tie-breaker for identical timestamps
                    same_date_rows['orig_position'] = range(len(same_date_rows))
                    
                    # Sort by time_int first, then by original position
                    same_date_rows = same_date_rows.sort_values(['time_int', 'orig_position'])
                    
                    # Add a rank column (1-based)
                    same_date_rows['rank'] = range(1, len(same_date_rows) + 1)
                    
                    # Find the rank of the current missing ID row
                    missing_row_in_sorted = same_date_rows[same_date_rows.index == idx]
                    
                    if not missing_row_in_sorted.empty:
                        rank = missing_row_in_sorted['rank'].iloc[0]
                        
                        # Create a new ID by combining the prefix with the rank (as 4 digits with leading zeros)
                        new_id = f"{id_prefix}{rank:04d}"
                        
                        results.append(f"Row {idx}: Reconstructed ID {new_id} from template/date with rank {rank}")
                        
                        # Update the ID in the dataframe
                        df.at[idx, id_col] = new_id
                        reconstructed_count += 1
                        continue
            
            # METHOD 2: If this is the only record for this date or METHOD 1 failed, 
            # directly extract date parts and create an ID
            id_prefix = extract_date_parts(reception_date)
            
            if id_prefix:
                # Since this is the only record or we couldn't use ranking, use 0001 as default
                new_id = f"{id_prefix}0001"
                results.append(f"Row {idx}: Created ID {new_id} directly from date {reception_date}")
                
                # Update the ID in the dataframe
                df.at[idx, id_col] = new_id
                solo_reconstructed_count += 1
            else:
                results.append(f"Row {idx}: Failed to extract date parts from {reception_date}")
        
        results.append(f"ID Reconstruction complete.")
        results.append(f"- IDs reconstructed using time ranking: {reconstructed_count}")
        results.append(f"- IDs reconstructed directly from date: {solo_reconstructed_count}")
        results.append(f"- Total reconstructed: {reconstructed_count + solo_reconstructed_count} of {len(missing_id_rows)}")
    
    # If we need to convert existing IDs from old to new format
    if old_format and datum_prijmu_col is not None:
        results.append("CONVERTING EXISTING IDs FROM OLD TO NEW FORMAT")
        
        # Get rows with valid IDs in old format
        valid_id_mask = df[id_col].notna() & (df[id_col] != '')
        old_format_ids = valid_id_mask & df[id_col].str.len().isin([8, 9])  # Allow for 8 or 9 digits
        
        if old_format_ids.sum() > 0:
            convert_count = 0
            
            for idx, row in df[old_format_ids].iterrows():
                old_id = row[id_col]
                reception_date = row[datum_prijmu_col]
                
                if pd.isna(reception_date):
                    results.append(f"Row {idx}: Cannot convert ID {old_id} - missing reception date")
                    continue
                
                # Extract new date prefix (YYMMDD)
                new_prefix = extract_date_parts(reception_date)
                
                if not new_prefix:
                    results.append(f"Row {idx}: Failed to extract date parts from {reception_date} for ID {old_id}")
                    continue
                
                # Extract sequence number from old ID (last 4 digits)
                seq_num = old_id[-4:]
                
                # Create new ID
                new_id = f"{new_prefix}{seq_num}"
                
                # Update ID
                df.at[idx, id_col] = new_id
                convert_count += 1
                
                # Print sample conversions (first 5)
                if convert_count <= 5:
                    results.append(f"Row {idx}: Converted {old_id} → {new_id}")
            
            results.append(f"Converted {convert_count} IDs from old format to new format.")
    
    # Find validovany vysledok column
    validovany_col = None
    for col in df.columns:
        if 'validovany' in str(col).lower() and 'vysledok' in str(col).lower():
            validovany_col = col
            break
    
    # Remove rows with blank validovany vysledok
    if validovany_col and remove_blank_validovany:
        initial_rows = len(df)
        df = df[df[validovany_col].notna() & (df[validovany_col] != '')]
        removed_rows = initial_rows - len(df)
        results.append(f"Removed {removed_rows} rows with blank values in '{validovany_col}' column")
    
    # Find diagnoza column
    diagnoza_col = None
    for col in df.columns:
        if 'diagnoza' in str(col).lower() or 'mkch' in str(col).lower():
            diagnoza_col = col
            break
    
    # Remove rows with blank diagnoza
    if diagnoza_col and remove_blank_diagnoza:
        initial_rows = len(df)
        df = df[df[diagnoza_col].notna() & (df[diagnoza_col] != '')]
        removed_rows = initial_rows - len(df)
        results.append(f"Removed {removed_rows} rows with blank values in '{diagnoza_col}' column")
    
    # Find HFE columns
    hfe_columns = []
    for col in df.columns:
        if 'hfe' in str(col).lower():
            hfe_columns.append(col)
    
    # Remove rows with blank HFE values
    if hfe_columns and remove_blank_hfe:
        initial_rows = len(df)
        for hfe_col in hfe_columns:
            df = df[df[hfe_col].notna() & (df[hfe_col] != '')]
        removed_rows = initial_rows - len(df)
        results.append(f"Removed {removed_rows} rows with blank values in HFE columns: {', '.join(hfe_columns)}")
    
    results.append(f"Final rows after cleaning: {len(df)}")
    return df, results

def analyze_dataset(df):
    """
    Perform basic analysis on the dataset and return results
    """
    results = []
    results.append(f"Dataset Overview:")
    results.append(f"Number of rows: {df.shape[0]}")
    results.append(f"Number of columns: {df.shape[1]}")
    
    # Column names and indices
    results.append("\nColumn Names and Indices:")
    for i, col in enumerate(df.columns):
        results.append(f"Column {i}: '{col}'")
    
    # Missing values analysis
    results.append("\nMissing Values Analysis:")
    missing = df.isnull().sum()
    missing_percent = (missing / len(df)) * 100
    missing_data = pd.DataFrame({
        'Column': missing.index,
        'Missing Values': missing.values,
        'Percentage (%)': missing_percent.values
    })
    missing_data = missing_data[missing_data['Missing Values'] > 0].reset_index(drop=True)
    if not missing_data.empty:
        for i, row in missing_data.iterrows():
            results.append(f"{row['Column']}: {row['Missing Values']} ({row['Percentage (%)']:.2f}%)")
    else:
        results.append("No missing values found")
    
    # Find HFE columns
    hfe_columns = []
    for col in df.columns:
        if 'hfe' in str(col).lower():
            hfe_columns.append(col)
    
    # Analyze HFE mutation distributions
    if hfe_columns:
        results.append("\nHFE Mutation Distribution:")
        for column in hfe_columns:
            results.append(f"\n{column}:")
            genotype_counts = df[column].value_counts()
            genotype_percentages = (genotype_counts / len(df) * 100).round(2)
            for genotype, count in genotype_counts.items():
                percentage = genotype_percentages[genotype]
                results.append(f"{genotype}: {count} patients ({percentage:.2f}%)")
    
    return results

def classify_hh_risk(row):
    """
    Classify hereditary hemochromatosis risk based on genotype combinations
    """
    # Extract genotypes, handling possible different column names
    c282y = h63d = s65c = None
    
    # Find the columns containing each mutation
    for col in row.index:
        if 'C282Y' in col:
            c282y = row[col]
        elif 'H63D' in col:
            h63d = row[col]
        elif 'S65C' in col:
            s65c = row[col]
    
    if not all([c282y, h63d, s65c]):
        return "Unknown (missing data)"
    
    # Check for C282Y homozygotes (highest risk)
    if c282y == "mutant":
        return "High Risk (C282Y homozygote)"
    
    # Check for compound heterozygotes (moderate risk)
    if c282y == "heterozygot" and h63d == "heterozygot":
        return "Moderate Risk (C282Y/H63D compound heterozygote)"
    if c282y == "heterozygot" and s65c == "heterozygot":
        return "Moderate Risk (C282Y/S65C compound heterozygote)"
    if h63d == "heterozygot" and s65c == "heterozygot":
        return "Lower Risk (H63D/S65C compound heterozygote)"
    
    # Check for H63D homozygotes (low to moderate risk)
    if h63d == "mutant":
        return "Lower Risk (H63D homozygote)"
    
    # Check for S65C homozygotes (rare, but considered lower risk)
    if s65c == "mutant":
        return "Lower Risk (S65C homozygote)"
    
    # Check for carriers (lower risk)
    if c282y == "heterozygot":
        return "Carrier (C282Y heterozygote)"
    if h63d == "heterozygot":
        return "Carrier (H63D heterozygote)"
    if s65c == "heterozygot":
        return "Carrier (S65C heterozygote)"
    
    # Everyone else
    return "Minimal Risk (no mutations)"

def analyze_hh_risk(df, hfe_columns):
    """
    Analyze hereditary hemochromatosis risk in the dataset
    """
    results = []
    results.append("HEREDITARY HEMOCHROMATOSIS RISK ANALYSIS")
    
    # Classify each patient's risk
    df["HH_Risk"] = df.apply(classify_hh_risk, axis=1)
    
    # Create a simplified risk category for statistical analysis
    def simplify_risk(risk):
        if "High Risk" in risk or "Moderate Risk" in risk:
            return "High/Moderate Risk"
        elif "Lower Risk" in risk:
            return "Lower Risk"
        elif "Carrier" in risk:
            return "Carrier"
        else:
            return "Minimal Risk"
    
    df["Risk_Category"] = df["HH_Risk"].apply(simplify_risk)
    
    # Count and percentage for each risk category
    risk_counts = df["HH_Risk"].value_counts()
    risk_percentages = (risk_counts / len(df) * 100).round(2)
    
    results.append("\nRisk Category Distribution:")
    for risk_category, count in risk_counts.items():
        results.append(f"{risk_category}: {count} patients ({risk_percentages[risk_category]:.2f}%)")
    
    # Summarize carriers and at-risk individuals
    carriers = df["HH_Risk"].str.contains("Carrier").sum()
    high_risk = df["HH_Risk"].str.contains("High Risk").sum()
    moderate_risk = df["HH_Risk"].str.contains("Moderate Risk").sum()
    lower_risk = df["HH_Risk"].str.contains("Lower Risk").sum()
    
    total_with_predisposition = high_risk + moderate_risk + lower_risk
    
    results.append("\nSUMMARY:")
    results.append(f"Total Carriers: {carriers} patients ({carriers/len(df)*100:.2f}%)")
    results.append(f"Total with Genetic Predisposition: {total_with_predisposition} patients ({total_with_predisposition/len(df)*100:.2f}%)")
    results.append(f"- High Risk: {high_risk} patients ({high_risk/len(df)*100:.2f}%)")
    results.append(f"- Moderate Risk: {moderate_risk} patients ({moderate_risk/len(df)*100:.2f}%)")
    results.append(f"- Lower Risk: {lower_risk} patients ({lower_risk/len(df)*100:.2f}%)")
    
    return df, results

def analyze_diagnosis_associations(df, hfe_columns):
    """
    Analyze associations between HFE mutations and diagnoses
    """
    results = []
    results.append("HFE MUTATIONS AND DIAGNOSIS ASSOCIATION ANALYSIS")
    
    # Define diagnosis categories of interest
    liver_disease_codes = ['K76.0', 'K75.9', 'K70', 'K71', 'K72', 'K73', 'K74', 'K76', 'K77']
    
    # Create a helper function to categorize diagnoses
    def categorize_diagnosis(diagnosis):
        if pd.isna(diagnosis):
            return "Unknown"
        
        # Check for liver diseases of interest
        for code in liver_disease_codes:
            if str(diagnosis).startswith(code):
                return "Liver Disease"
        
        # Check for other major categories
        if str(diagnosis).startswith('K'):
            return "Other Digestive System"
        elif str(diagnosis).startswith('E'):
            return "Endocrine/Metabolic"
        elif str(diagnosis).startswith('B'):
            return "Infectious Disease"
        else:
            return "Other"
    
    # Find diagnoza column
    diagnoza_col = None
    for col in df.columns:
        if 'diagnoza' in str(col).lower() or 'mkch' in str(col).lower():
            diagnoza_col = col
            break
    
    if not diagnoza_col:
        results.append("No diagnosis column found in the dataset")
        return df, results
    
    # Add diagnosis category column
    df['Diagnosis_Category'] = df[diagnoza_col].apply(categorize_diagnosis)
    
    # Also create a specific column for our two main liver diseases of interest
    def is_specific_liver_disease(diagnosis):
        if pd.isna(diagnosis):
            return "Other"
        if str(diagnosis).startswith('K76.0'):
            return "K76.0 (Fatty liver)"
        elif str(diagnosis).startswith('K75.9'):
            return "K75.9 (Inflammatory liver disease)"
        else:
            return "Other"
    
    df['Specific_Liver_Disease'] = df[diagnoza_col].apply(is_specific_liver_disease)
    
    # Create a column for any HFE mutation
    def has_any_mutation(row):
        for col in hfe_columns:
            if row[col] != 'normal':
                return "Mutation Present"
        return "No Mutation"
        
    df['Any_HFE_Mutation'] = df.apply(has_any_mutation, axis=1)
    
    # Print basic statistics
    results.append("\nDiagnosis category distribution:")
    diag_dist = df['Diagnosis_Category'].value_counts()
    for category, count in diag_dist.items():
        results.append(f"{category}: {count} patients ({count/len(df)*100:.2f}%)")
    
    # Create a binary column for each specific liver disease
    df['Has_K760'] = df[diagnoza_col].apply(lambda x: 1 if str(x).startswith('K76.0') else 0)
    df['Has_K759'] = df[diagnoza_col].apply(lambda x: 1 if str(x).startswith('K75.9') else 0)
    
    # Calculate prevalence in patients with and without mutations
    k760_in_mutation = df[df['Any_HFE_Mutation'] == 'Mutation Present']['Has_K760'].mean() * 100
    k760_in_no_mutation = df[df['Any_HFE_Mutation'] == 'No Mutation']['Has_K760'].mean() * 100
    
    k759_in_mutation = df[df['Any_HFE_Mutation'] == 'Mutation Present']['Has_K759'].mean() * 100
    k759_in_no_mutation = df[df['Any_HFE_Mutation'] == 'No Mutation']['Has_K759'].mean() * 100
    
    results.append(f"\nK76.0 (Fatty liver) prevalence:")
    results.append(f"- In patients with HFE mutations: {k760_in_mutation:.2f}%")
    results.append(f"- In patients without HFE mutations: {k760_in_no_mutation:.2f}%")
    
    results.append(f"\nK75.9 (Inflammatory liver disease) prevalence:")
    results.append(f"- In patients with HFE mutations: {k759_in_mutation:.2f}%")
    results.append(f"- In patients without HFE mutations: {k759_in_no_mutation:.2f}%")
    
    # Create a 2x2 contingency table for each specific disease
    k760_table = pd.crosstab(df['Any_HFE_Mutation'], df['Has_K760'])
    k759_table = pd.crosstab(df['Any_HFE_Mutation'], df['Has_K759'])
    
    # Perform chi-square tests on the 2x2 tables if possible
    try:
        k760_chi2, k760_p, k760_dof, k760_expected = chi2_contingency(k760_table)
        results.append(f"\nChi-Square Test for K76.0: chi2={k760_chi2:.2f}, p={k760_p:.4f}")
        if k760_p < 0.05:
            results.append("There is a significant association between HFE mutations and K76.0 (Fatty liver).")
        else:
            results.append("No significant association found between HFE mutations and K76.0 (Fatty liver).")
    except:
        results.append("Could not perform chi-square test for K76.0 - may have insufficient data.")
    
    try:
        k759_chi2, k759_p, k759_dof, k759_expected = chi2_contingency(k759_table)
        results.append(f"\nChi-Square Test for K75.9: chi2={k759_chi2:.2f}, p={k759_p:.4f}")
        if k759_p < 0.05:
            results.append("There is a significant association between HFE mutations and K75.9 (Inflammatory liver disease).")
        else:
            results.append("No significant association found between HFE mutations and K75.9 (Inflammatory liver disease).")
    except:
        results.append("Could not perform chi-square test for K75.9 - may have insufficient data.")
    
    return df, results

def generate_genotype_distribution_plot(df, hfe_columns):
    """
    Generate plots for genotype distribution
    """
    plots = []
    
    # Create a plot for each HFE column
    for column in hfe_columns:
        plt.figure(figsize=(10, 6))
        genotype_counts = df[column].value_counts()
        total = len(df)
        
        # Create bar chart
        ax = sns.barplot(x=genotype_counts.index, y=genotype_counts.values)
        
        # Add count and percentage labels
        for i, p in enumerate(ax.patches):
            height = p.get_height()
            percentage = height / total * 100
            ax.annotate(f'{int(height)} ({percentage:.1f}%)', 
                        (p.get_x() + p.get_width() / 2., height),
                        ha='center', va='bottom', xytext=(0, 10),
                        textcoords='offset points')
        
        # Add labels and title
        plt.xlabel('Genotype')
        plt.ylabel('Number of Patients')
        column_name = column.split('\n')[0] if '\n' in column else column
        plt.title(f'Genotype Distribution for {column_name}')
        plt.tight_layout()
        
        # Convert to base64 for display in Shiny
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        img_str = base64.b64encode(buf.read()).decode('utf-8')
        plots.append((column_name, img_str))
        plt.close()
    
    return plots

def generate_risk_distribution_plot(df):
    """
    Generate plot for HH risk distribution
    """
    if 'Risk_Category' not in df.columns:
        return None
    
    plt.figure(figsize=(10, 6))
    risk_counts = df['Risk_Category'].value_counts()
    total = len(df)
    
    # Create bar chart
    ax = sns.barplot(x=risk_counts.index, y=risk_counts.values)
    
    # Add count and percentage labels
    for i, p in enumerate(ax.patches):
        height = p.get_height()
        percentage = height / total * 100
        ax.annotate(f'{int(height)} ({percentage:.1f}%)', 
                    (p.get_x() + p.get_width() / 2., height),
                    ha='center', va='bottom', xytext=(0, 10),
                    textcoords='offset points')
    
    # Add labels and title
    plt.xlabel('Risk Category')
    plt.ylabel('Number of Patients')
    plt.title('Hereditary Hemochromatosis Risk Distribution')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    
    # Convert to base64 for display in Shiny
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    img_str = base64.b64encode(buf.read()).decode('utf-8')
    plt.close()
    
    return img_str

def generate_diagnosis_association_plot(df):
    """
    Generate plot for diagnosis association with HFE mutations
    """
    if 'Diagnosis_Category' not in df.columns or 'Any_HFE_Mutation' not in df.columns:
        return None
    
    plt.figure(figsize=(12, 8))
    diagnosis_mutation = pd.crosstab(df['Diagnosis_Category'], df['Any_HFE_Mutation'], normalize='index') * 100
    diagnosis_mutation.plot(kind='bar', stacked=True)
    plt.title('Diagnosis Categories by HFE Mutation Status')
    plt.xlabel('Diagnosis Category')
    plt.ylabel('Percentage (%)')
    plt.xticks(rotation=45, ha='right')
    plt.legend(title='Mutation Status')
    plt.tight_layout()
    
    # Convert to base64 for display in Shiny
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    img_str = base64.b64encode(buf.read()).decode('utf-8')
    plt.close()
    
    return img_str

def generate_liver_disease_plot(df):
    """
    Generate plot for liver disease prevalence by risk category
    """
    if 'Risk_Category' not in df.columns or 'Has_K760' not in df.columns or 'Has_K759' not in df.columns:
        return None
    
    k760_by_risk = df.groupby('Risk_Category')['Has_K760'].mean() * 100
    k759_by_risk = df.groupby('Risk_Category')['Has_K759'].mean() * 100
    
    liver_prevalence = pd.DataFrame({
        'K76.0 (Fatty liver)': k760_by_risk,
        'K75.9 (Inflammatory liver disease)': k759_by_risk
    })
    
    plt.figure(figsize=(12, 8))
    liver_prevalence.plot(kind='bar')
    plt.title('Liver Disease Prevalence by HH Risk Category')
    plt.xlabel('Risk Category')
    plt.ylabel('Prevalence (%)')
    plt.xticks(rotation=45, ha='right')
    plt.legend(title='Liver Disease')
    plt.tight_layout()
    
    # Convert to base64 for display in Shiny
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    img_str = base64.b64encode(buf.read()).decode('utf-8')
    plt.close()
    
    return img_str

def calculate_hwe(df, mutation_column, column_name):
    """
    Calculate Hardy-Weinberg equilibrium for a given mutation column.
    
    Args:
        df: DataFrame containing the genotype data
        mutation_column: Column name containing genotype data
        column_name: Human-readable name for the mutation
        
    Returns:
        Dictionary with HWE analysis results
    """
    # Count genotypes
    genotype_counts = df[mutation_column].value_counts()
    
    # Extract counts for each genotype (handle case where a genotype might be missing)
    normal_count = genotype_counts.get('normal', 0)
    heterozygote_count = genotype_counts.get('heterozygot', 0)
    mutant_count = genotype_counts.get('mutant', 0)
    
    total = normal_count + heterozygote_count + mutant_count
    
    # Calculate allele frequencies
    p = (2 * normal_count + heterozygote_count) / (2 * total)  # Normal allele frequency
    q = (2 * mutant_count + heterozygote_count) / (2 * total)  # Mutant allele frequency
    
    # Calculate expected genotype counts under Hardy-Weinberg equilibrium
    expected_normal = (p**2) * total
    expected_heterozygote = 2 * p * q * total
    expected_mutant = (q**2) * total
    
    # Perform chi-square test
    observed = np.array([normal_count, heterozygote_count, mutant_count])
    expected = np.array([expected_normal, expected_heterozygote, expected_mutant])
    
    # Filter out zeros to avoid division by zero in chi-square test
    valid_indices = expected > 0
    
    if sum(valid_indices) <= 1:
        # Not enough valid categories for chi-square test
        chi2 = np.nan
        p_value = np.nan
        df_freedom = 0
        is_in_hwe = "Cannot determine (insufficient data)"
    else:
        # Calculate chi-square and p-value using scipy.stats.chisquare
        from scipy import stats
        chi2, p_value = stats.chisquare(
            observed[valid_indices], 
            expected[valid_indices]
        )
        
        # Degrees of freedom: number of categories - 1 (allele frequency) - 1 = #categories - 2
        df_freedom = sum(valid_indices) - 1 - 1
        df_freedom = max(1, df_freedom)  # Ensure at least 1 degree of freedom
        
        # Determine if in Hardy-Weinberg equilibrium
        is_in_hwe = "Yes" if p_value > 0.05 else "No"
    
    # Return results
    results = {
        "Mutation": column_name,
        "Total": total,
        "Normal (observed)": normal_count,
        "Heterozygote (observed)": heterozygote_count,
        "Mutant (observed)": mutant_count,
        "Normal allele frequency (p)": p,
        "Mutant allele frequency (q)": q,
        "p + q": p + q,
        "Normal (expected)": expected_normal,
        "Heterozygote (expected)": expected_heterozygote,
        "Mutant (expected)": expected_mutant,
        "Chi-square": chi2,
        "p-value": p_value,
        "Degrees of freedom": df_freedom,
        "In Hardy-Weinberg equilibrium": is_in_hwe
    }
    
    return results

def analyze_hardy_weinberg(df, hfe_columns):
    """
    Analyze Hardy-Weinberg equilibrium for HFE mutations in the dataset
    """
    results = []
    results.append("HARDY-WEINBERG EQUILIBRIUM ANALYSIS")
    
    # Define human-readable names for mutations
    mutation_names = {
        "HFE C187G (H63D)\n[HFE]": "HFE H63D",
        "HFE A193T (S65C)\n[HFE]": "HFE S65C", 
        "HFE G845A (C282Y)\n[HFE]": "HFE C282Y"
    }
    
    # Store calculated HWE results for plotting
    hwe_results_list = []
    
    # Run Hardy-Weinberg equilibrium analysis for each mutation
    for column in hfe_columns:
        # Get human-readable name if available, otherwise use column name
        column_name = mutation_names.get(column, column)
        
        results.append(f"\n{'-'*50}")
        results.append(f"Analyzing {column_name}")
        results.append(f"{'-'*50}")
        
        # Print genotype distribution
        genotype_counts = df[column].value_counts()
        results.append("\nGenotype Distribution:")
        for genotype, count in genotype_counts.items():
            percentage = count / len(df) * 100
            results.append(f"{genotype}: {count} ({percentage:.2f}%)")
        
        # Calculate Hardy-Weinberg equilibrium
        hwe_results = calculate_hwe(df, column, column_name)
        hwe_results_list.append(hwe_results)
        
        # Print results
        results.append("\nHardy-Weinberg Equilibrium Analysis:")
        results.append(f"Total individuals: {hwe_results['Total']}")
        results.append(f"Normal allele frequency (p): {hwe_results['Normal allele frequency (p)']:.4f}")
        results.append(f"Mutant allele frequency (q): {hwe_results['Mutant allele frequency (q)']:.4f}")
        results.append(f"Sum (p + q): {hwe_results['p + q']:.4f} (should be close to 1.0)")
        
        results.append("\nObserved vs. Expected Counts:")
        results.append(f"Normal (wildtype): {hwe_results['Normal (observed)']:.1f} observed vs {hwe_results['Normal (expected)']:.1f} expected")
        results.append(f"Heterozygote: {hwe_results['Heterozygote (observed)']:.1f} observed vs {hwe_results['Heterozygote (expected)']:.1f} expected")
        results.append(f"Mutant (homozygote): {hwe_results['Mutant (observed)']:.1f} observed vs {hwe_results['Mutant (expected)']:.1f} expected")
        
        results.append("\nStatistical Test:")
        if np.isnan(hwe_results['Chi-square']):
            results.append("Chi-square test could not be performed (insufficient data)")
        else:
            results.append(f"Chi-square: {hwe_results['Chi-square']:.4f}")
            results.append(f"p-value: {hwe_results['p-value']:.4f}")
            results.append(f"Degrees of freedom: {hwe_results['Degrees of freedom']}")
        
        results.append(f"\nIn Hardy-Weinberg equilibrium: {hwe_results['In Hardy-Weinberg equilibrium']}")
    
    return df, results, hwe_results_list

def generate_hardy_weinberg_plots(hwe_results_list):
    """
    Create bar charts comparing observed vs expected genotype distributions for HWE analysis
    """
    plots = []
    
    for result in hwe_results_list:
        mutation = result["Mutation"]
        
        # Data for plotting
        genotypes = ["Normal", "Heterozygote", "Mutant"]
        observed = [result["Normal (observed)"], result["Heterozygote (observed)"], result["Mutant (observed)"]]
        expected = [result["Normal (expected)"], result["Heterozygote (expected)"], result["Mutant (expected)"]]
        
        # Create figure
        plt.figure(figsize=(10, 6))
        
        # Set up bar width and positions
        bar_width = 0.35
        indices = np.arange(len(genotypes))
        
        # Create bars
        plt.bar(indices - bar_width/2, observed, bar_width, label='Observed', color='blue', alpha=0.7)
        plt.bar(indices + bar_width/2, expected, bar_width, label='Expected (HWE)', color='green', alpha=0.7)
        
        # Add labels, title, and legend
        plt.xlabel('Genotype')
        plt.ylabel('Count')
        plt.title(f'Hardy-Weinberg Equilibrium Analysis: {mutation}')
        plt.xticks(indices, genotypes)
        plt.legend()
        
        # Add HWE status and p-value to the plot
        hwe_status = result["In Hardy-Weinberg equilibrium"]
        p_value = result["p-value"]
        
        if isinstance(p_value, float):
            plt.figtext(0.5, 0.01, f"HWE Status: {hwe_status} (p = {p_value:.4f})", 
                        ha="center", fontsize=12, bbox={"facecolor":"orange", "alpha":0.5, "pad":5})
        else:
            plt.figtext(0.5, 0.01, f"HWE Status: {hwe_status}", 
                        ha="center", fontsize=12, bbox={"facecolor":"orange", "alpha":0.5, "pad":5})
        
        # Convert to base64 for display in Shiny
        buf = io.BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight')
        buf.seek(0)
        img_str = base64.b64encode(buf.read()).decode('utf-8')
        plots.append((mutation, img_str))
        plt.close()
    
    return plots

def generate_genotype_by_age_plot(df, hfe_columns):
    """Generate plots showing the relationship between genotypes and age"""
    plots = []
    results = []
    
    # Find age column - expanded search terms
    age_col = next((col for col in df.columns if any(term in str(col).lower() for term in 
                   ['vek', 'age', 'rok', 'year', 'date', 'datum', 'nar', 'birth'])), None)
    
    results.append(f"Age column search: {'Found: ' + age_col if age_col else 'Not found'}")
    
    if age_col is None:
        # Try to infer from date columns
        date_cols = [col for col in df.columns if any(term in str(col).lower() for term in 
                    ['date', 'datum', 'cas', 'time'])]
        results.append(f"Checked date columns: {date_cols}")
        return [], results
    
    # Print sample values from age column for diagnostics
    sample_values = df[age_col].dropna().head(5).tolist()
    results.append(f"Sample values in age column: {sample_values}")
    
    # Ensure age column is numeric
    try:
        # If it's a date column, extract year
        if pd.api.types.is_datetime64_any_dtype(df[age_col]):
            results.append(f"Detected datetime in age column")
            current_year = datetime.now().year
            df['calculated_age'] = current_year - df[age_col].dt.year
            age_col = 'calculated_age'
            results.append(f"Created calculated age based on year")
        else:
            # Try to convert to numeric
            df[age_col] = pd.to_numeric(df[age_col], errors='coerce')
        
        # Check if conversion worked
        if df[age_col].isna().all():
            results.append(f"Age column contains no valid numeric values after conversion")
            return [], results
            
        # Check the range of ages to see if they're reasonable
        min_age = df[age_col].min()
        max_age = df[age_col].max()
        results.append(f"Age range: {min_age} to {max_age}")
        
        # If "ages" are very large (e.g., years like 1990), they might be birth years
        if min_age > 120:  # Assuming no one is older than 120
            current_year = datetime.now().year
            df['calculated_age'] = current_year - df[age_col]
            age_col = 'calculated_age'
            results.append(f"Converted birth years to ages. New range: {df[age_col].min()} to {df[age_col].max()}")
    except Exception as e:
        results.append(f"Error converting age column: {str(e)}")
        return [], results
    
    # Create age groups for better visualization
    try:
        df['Age_Group'] = pd.cut(df[age_col], bins=[0, 30, 45, 60, 75, 100], 
                                labels=['<30', '30-45', '45-60', '60-75', '>75'])
        results.append(f"Created age groups successfully")
    except Exception as e:
        results.append(f"Error creating standard age groups: {str(e)}")
        try:
            # Alternative approach with quantiles
            df['Age_Group'] = pd.qcut(df[age_col].dropna(), q=4, duplicates='drop')
            results.append(f"Created quartile-based age groups")
        except Exception as e2:
            results.append(f"Error creating alternative age groups: {str(e2)}")
            return [], results
    
    # Count how many individuals in each age group
    age_group_counts = df['Age_Group'].value_counts()
    results.append(f"Age group distribution: {dict(age_group_counts)}")
    
    for column in hfe_columns:
        try:
            plt.figure(figsize=(12, 8))
            
            # Filter out NaN values in both columns
            valid_data = df.dropna(subset=[column, 'Age_Group'])
            results.append(f"Valid data points for {column}: {len(valid_data)}")
            
            # Skip if not enough data (reduced threshold)
            if len(valid_data) < 3:
                results.append(f"Not enough data for {column}")
                plt.close()
                continue
                
            mutation_by_age = pd.crosstab(valid_data['Age_Group'], valid_data[column], normalize='index') * 100
            
            # Skip if crosstab is empty
            if mutation_by_age.empty:
                results.append(f"Empty crosstab for {column}")
                plt.close()
                continue
                
            mutation_by_age.plot(kind='bar', stacked=True)
            
            plt.title(f'Age Distribution by {column} Genotype')
            plt.xlabel('Age Group')
            plt.ylabel('Percentage (%)')
            plt.legend(title='Genotype')
            plt.tight_layout()
            
            # Convert to base64
            buf = io.BytesIO()
            plt.savefig(buf, format='png')
            buf.seek(0)
            img_str = base64.b64encode(buf.read()).decode('utf-8')
            plots.append((f"{column} by Age", img_str))
            plt.close()
            results.append(f"Successfully created plot for {column}")
        except Exception as e:
            results.append(f"Error creating plot for {column}: {str(e)}")
            plt.close()
            continue
        
    return plots, results

def generate_genotype_by_gender_plot(df, hfe_columns):
    """Generate plots showing the relationship between genotypes and gender"""
    plots = []
    results = []
    
    # Find gender column - expanded search terms
    gender_col = next((col for col in df.columns 
                      if any(term in str(col).lower() for term in 
                            ['pohlavie', 'gender', 'sex', 'pohlavi', 'sexus', 'm/f', 'm/ž', 'm/z'])), None)
    
    results.append(f"Gender column search: {'Found: ' + gender_col if gender_col else 'Not found'}")
    
    if gender_col is None:
        # Check for columns with short values that might be gender
        short_value_cols = []
        for col in df.columns:
            # Check if column contains mostly short strings (potential gender values)
            if df[col].dtype == object:
                val_lengths = df[col].astype(str).str.len()
                if val_lengths.mean() < 3 and val_lengths.max() < 5:
                    short_value_cols.append(col)
        
        results.append(f"Potential gender columns based on value length: {short_value_cols}")
        
        # If no gender column found, try the first short value column
        if short_value_cols and not gender_col:
            gender_col = short_value_cols[0]
            results.append(f"Using column {gender_col} as potential gender column")
    
    if gender_col is None:
        return [], results
    
    # Print sample values from gender column for diagnostics
    sample_values = df[gender_col].dropna().head(5).tolist()
    results.append(f"Sample values in gender column: {sample_values}")
    
    # Standardize gender values
    try:
        # Create a copy to avoid modifying the original DataFrame
        gender_mapping = {}
        unique_values = df[gender_col].dropna().unique()
        
        for val in unique_values:
            str_val = str(val).lower().strip()
            if str_val in ['m', 'male', 'muz', 'muž', '1']:
                gender_mapping[val] = 'Male'
            elif str_val in ['f', 'female', 'zena', 'žena', 'z', '2']:
                gender_mapping[val] = 'Female'
                
        # If we found mappings, create a standardized gender column
        if gender_mapping:
            df['Standardized_Gender'] = df[gender_col].map(gender_mapping)
            gender_col = 'Standardized_Gender'
            results.append(f"Standardized gender values: {gender_mapping}")
        
        # Check the number of unique gender values
        gender_counts = df[gender_col].value_counts()
        results.append(f"Gender distribution: {dict(gender_counts)}")
    except Exception as e:
        results.append(f"Error standardizing gender: {str(e)}")
    
    for column in hfe_columns:
        try:
            plt.figure(figsize=(10, 6))
            
            # Filter out NaN values in both columns
            valid_data = df.dropna(subset=[column, gender_col])
            results.append(f"Valid data points for {column}: {len(valid_data)}")
            
            # Skip if not enough data (reduced threshold)
            if len(valid_data) < 3:
                results.append(f"Not enough data for {column}")
                plt.close()
                continue
                
            gender_by_mutation = pd.crosstab(valid_data[gender_col], valid_data[column], normalize='index') * 100
            
            # Skip if crosstab is empty
            if gender_by_mutation.empty:
                results.append(f"Empty crosstab for {column}")
                plt.close()
                continue
                
            gender_by_mutation.plot(kind='bar', stacked=True)
            
            plt.title(f'Gender Distribution by {column} Genotype')
            plt.xlabel('Gender')
            plt.ylabel('Percentage (%)')
            plt.legend(title='Genotype')
            plt.tight_layout()
            
            # Convert to base64
            buf = io.BytesIO()
            plt.savefig(buf, format='png')
            buf.seek(0)
            img_str = base64.b64encode(buf.read()).decode('utf-8')
            plots.append((f"{column} by Gender", img_str))
            plt.close()
            results.append(f"Successfully created plot for {column}")
        except Exception as e:
            results.append(f"Error creating plot for {column}: {str(e)}")
            plt.close()
            continue
        
    return plots, results

def generate_genotype_diagnosis_gender_plot(df, hfe_columns):
    """Generate plots showing relationships between genotypes, diagnoses, and gender"""
    plots = []
    results = []
    
    # Find gender and diagnosis columns - expanded search terms
    gender_col = next((col for col in df.columns 
                      if any(term in str(col).lower() for term in 
                            ['pohlavie', 'gender', 'sex', 'pohlavi', 'sexus', 'm/f', 'm/ž', 'm/z'])), None)
    
    diagnoza_col = next((col for col in df.columns 
                        if any(term in str(col).lower() for term in 
                              ['diagnoza', 'mkch', 'diagnosis', 'icd', 'kod'])), None)
    
    results.append(f"Gender column search: {'Found: ' + gender_col if gender_col else 'Not found'}")
    results.append(f"Diagnosis column search: {'Found: ' + diagnoza_col if diagnoza_col else 'Not found'}")
    
    if gender_col is None:
        # Check for columns with short values that might be gender
        short_value_cols = []
        for col in df.columns:
            # Check if column contains mostly short strings (potential gender values)
            if df[col].dtype == object:
                val_lengths = df[col].astype(str).str.len()
                if val_lengths.mean() < 3 and val_lengths.max() < 5:
                    short_value_cols.append(col)
        
        results.append(f"Potential gender columns based on value length: {short_value_cols}")
        
        # If no gender column found, try the first short value column
        if short_value_cols and not gender_col:
            gender_col = short_value_cols[0]
            results.append(f"Using column {gender_col} as potential gender column")
    
    if gender_col is None or diagnoza_col is None:
        if gender_col is None:
            results.append("Cannot create gender-diagnosis plots: gender column not found")
        if diagnoza_col is None:
            results.append("Cannot create gender-diagnosis plots: diagnosis column not found")
        return [], results
    
    # Print sample values for diagnostics
    gender_samples = df[gender_col].dropna().head(5).tolist()
    diag_samples = df[diagnoza_col].dropna().head(5).tolist()
    results.append(f"Sample gender values: {gender_samples}")
    results.append(f"Sample diagnosis values: {diag_samples}")
    
    # Standardize gender values
    try:
        # Create a copy to avoid modifying the original DataFrame
        gender_mapping = {}
        unique_values = df[gender_col].dropna().unique()
        
        for val in unique_values:
            str_val = str(val).lower().strip()
            if str_val in ['m', 'male', 'muz', 'muž', '1']:
                gender_mapping[val] = 'Male'
            elif str_val in ['f', 'female', 'zena', 'žena', 'z', '2']:
                gender_mapping[val] = 'Female'
                
        # If we found mappings, create a standardized gender column
        if gender_mapping:
            df['Standardized_Gender'] = df[gender_col].map(gender_mapping)
            gender_col = 'Standardized_Gender'
            results.append(f"Standardized gender values: {gender_mapping}")
        
        # Check the number of unique gender values
        gender_counts = df[gender_col].value_counts()
        results.append(f"Gender distribution: {dict(gender_counts)}")
    except Exception as e:
        results.append(f"Error standardizing gender: {str(e)}")
    
    # Create diagnosis categories if not already there
    if 'Diagnosis_Category' not in df.columns:
        try:
            def categorize_diagnosis(diagnosis):
                if pd.isna(diagnosis):
                    return "Unknown"
                
                # Check for liver diseases
                liver_disease_codes = ['K76.0', 'K75.9', 'K70', 'K71', 'K72', 'K73', 'K74', 'K76', 'K77']
                for code in liver_disease_codes:
                    if str(diagnosis).startswith(code):
                        return "Liver Disease"
                
                # Check for other major categories
                if str(diagnosis).startswith('K'):
                    return "Other Digestive System"
                elif str(diagnosis).startswith('E'):
                    return "Endocrine/Metabolic"
                elif str(diagnosis).startswith('B'):
                    return "Infectious Disease"
                elif str(diagnosis).startswith('D'):
                    return "Blood Disorders"
                elif str(diagnosis).startswith('C'):
                    return "Neoplasms"
                else:
                    return "Other"
            
            df['Diagnosis_Category'] = df[diagnoza_col].apply(categorize_diagnosis)
            diag_cat_counts = df['Diagnosis_Category'].value_counts()
            results.append(f"Created diagnosis categories: {dict(diag_cat_counts)}")
        except Exception as e:
            results.append(f"Error creating diagnosis categories: {str(e)}")
            df['Diagnosis_Category'] = "Unknown"
    
    # For each HFE mutation
    for column in hfe_columns:
        # Create separate plots for each gender
        for gender_value in df[gender_col].dropna().unique():
            try:
                if pd.isna(gender_value):
                    continue
                    
                gender_subset = df[df[gender_col] == gender_value]
                results.append(f"Data for {gender_value}: {len(gender_subset)} rows")
                
                # Skip if too few samples (reduced threshold)
                if len(gender_subset) < 3:
                    results.append(f"Not enough data for {gender_value}")
                    continue
                
                # Filter out NaN values
                valid_data = gender_subset.dropna(subset=[column, 'Diagnosis_Category'])
                results.append(f"Valid data points for {column} and {gender_value}: {len(valid_data)}")
                
                # Skip if not enough data after filtering
                if len(valid_data) < 3:
                    results.append(f"Not enough valid data after filtering")
                    continue
                
                # Check if we have sufficient diagnosis categories
                diag_counts = valid_data['Diagnosis_Category'].value_counts()
                results.append(f"Diagnosis counts for {gender_value}: {dict(diag_counts)}")
                
                if len(diag_counts) < 2:
                    results.append(f"Not enough diagnosis categories for {gender_value}")
                    continue
                
                plt.figure(figsize=(12, 8))
                
                # Create crosstab 
                try:
                    diagnosis_by_mutation = pd.crosstab(
                        valid_data['Diagnosis_Category'], 
                        valid_data[column], 
                        normalize='index'
                    ) * 100
                    
                    # Skip if crosstab is empty or has only one row/column
                    if diagnosis_by_mutation.empty or diagnosis_by_mutation.shape[0] < 2 or diagnosis_by_mutation.shape[1] < 2:
                        results.append(f"Insufficient crosstab dimensions: {diagnosis_by_mutation.shape}")
                        plt.close()
                        continue
                        
                    diagnosis_by_mutation.plot(kind='bar', stacked=True)
                except Exception as e:
                    results.append(f"Error creating crosstab: {str(e)}")
                    plt.close()
                    continue
                
                plt.title(f'Diagnosis Distribution by {column} Genotype - {gender_value}')
                plt.xlabel('Diagnosis Category')
                plt.ylabel('Percentage (%)')
                plt.legend(title='Genotype')
                plt.xticks(rotation=45, ha='right')
                plt.tight_layout()
                
                # Convert to base64
                buf = io.BytesIO()
                plt.savefig(buf, format='png')
                buf.seek(0)
                img_str = base64.b64encode(buf.read()).decode('utf-8')
                plots.append((f"{column} by Diagnosis - {gender_value}", img_str))
                plt.close()
                results.append(f"Successfully created plot for {column} and {gender_value}")
            except Exception as e:
                results.append(f"Error creating plot for {column} and {gender_value}: {str(e)}")
                plt.close()
                continue
    
    return plots, results

# Define the Shiny UI
app_ui = ui.page_fluid(
    ui.h2("HFE Mutation Analysis"),
    
    ui.layout_sidebar(
        ui.sidebar(
            ui.h3("Upload Dataset"),
            ui.input_file("file1", "Choose Excel File", accept=[".xls", ".xlsx"]),
            ui.hr(),
            
            ui.h3("Data Cleaning Options"),
            ui.input_checkbox("clean_validovany", "Remove rows with blank validovany vysledok", True),
            ui.input_checkbox("clean_diagnoza", "Remove rows with blank diagnoza", True),
            ui.input_checkbox("clean_hfe", "Remove rows with blank HFE values", True),
            ui.input_checkbox("remove_second_col", "Remove second column (usually empty)", True),
            ui.input_action_button("btn_clean", "Clean Dataset", class_="btn-primary"),
            ui.output_ui("formatted_download_button"),
            ui.hr(),
            
            ui.h3("Analysis Options"),
            ui.input_radio_buttons(
                "analysis_type",
                "Analysis Type",
                {
                    "basic": "Basic Dataset Analysis",
                    "hh_risk": "HH Risk Analysis",
                    "diagnosis": "Diagnosis Association Analysis",
                    "hardy_weinberg": "Hardy-Weinberg Equilibrium Analysis",
                    "all": "Complete Analysis"
                },
                selected="all"
            ),
            ui.input_action_button("btn_analyze", "Run Analysis", class_="btn-success"),
            ui.hr(),
            
            ui.output_ui("data_download_button"),
        ),
        
        ui.navset_tab(
            ui.nav_panel("Dataset Overview",
                ui.h3("Dataset Information"),
                ui.output_text_verbatim("dataset_info"),
                ui.h3("Data Preview"),
                ui.output_data_frame("data_preview")
            ),
            ui.nav_panel("Analysis Results",
                ui.h3("Analysis Output"),
                ui.output_text_verbatim("analysis_results")
            ),
            ui.nav_panel("Visualizations",
                ui.h3("Genotype Distribution"),
                ui.output_ui("genotype_plots"),
                
                ui.h3("Genotypes by Age"),
                ui.output_ui("genotype_age_plots"),
                
                ui.h3("Genotypes by Gender"),
                ui.output_ui("genotype_gender_plots"),
                
                ui.h3("Genotypes by Diagnosis and Gender"),
                ui.output_ui("genotype_diagnosis_gender_plots"),
                
                ui.h3("Risk Distribution"),
                ui.output_ui("risk_plot"),
                ui.h3("Diagnosis Associations"),
                ui.output_ui("diagnosis_plot"),
                ui.h3("Liver Disease by Risk Category"),
                ui.output_ui("liver_disease_plot"),
                ui.h3("Hardy-Weinberg Equilibrium"),
                ui.output_ui("hardy_weinberg_plots")
            )
        )
    ),
    title="HFE Mutation Analysis",
    theme=shinyswatch.theme.superhero
)

# Define the Shiny server
def server(input, output, session):
    # Reactive value to store the dataset
    data = reactive.Value(None)
    data_cleaned = reactive.Value(False)
    analysis_run = reactive.Value(False)
    analysis_results_text = reactive.Value([])
    data_info = reactive.Value([])
    plots_genotype = reactive.Value([])
    plot_risk = reactive.Value(None)
    plot_diagnosis = reactive.Value(None)
    plot_liver = reactive.Value(None)
    plots_hardy_weinberg = reactive.Value([])
    
    # New reactive values for the new visualizations
    plots_age = reactive.Value([])
    plots_gender = reactive.Value([])
    plots_diagnosis_gender = reactive.Value([])
    
    # Add new diagnostic reactive values
    diagnostics_age = reactive.Value([])
    diagnostics_gender = reactive.Value([])
    diagnostics_diagnosis_gender = reactive.Value([])
    
    @reactive.Effect
    @reactive.event(input.file1)
    def _():
        if input.file1() is None:
            return
        
        file_info = input.file1()
        file_path = file_info[0]["datapath"]
        
        try:
            df = pd.read_excel(file_path)
            data.set(df)
            data_cleaned.set(False)
            analysis_run.set(False)
            
            # Basic dataset info
            info = []
            info.append(f"File: {file_info[0]['name']}")
            info.append(f"Rows: {len(df)}")
            info.append(f"Columns: {len(df.columns)}")
            
            # HFE columns
            hfe_columns = []
            for col in df.columns:
                if 'hfe' in str(col).lower():
                    hfe_columns.append(col)
            
            if hfe_columns:
                info.append(f"Found {len(hfe_columns)} HFE mutation columns:")
                for col in hfe_columns:
                    info.append(f"- {col}")
            else:
                info.append("No HFE mutation columns found in the dataset")
            
            data_info.set(info)
        except Exception as e:
            ui.notification_show(f"Error loading file: {str(e)}", type="error", duration=None)
    
    @reactive.Effect
    @reactive.event(input.btn_clean)
    def _():
        if data() is None:
            ui.notification_show("Please upload a dataset first", type="warning")
            return
        
        try:
            df, results = clean_dataset(
                data(), 
                remove_blank_validovany=input.clean_validovany(),
                remove_blank_diagnoza=input.clean_diagnoza(),
                remove_blank_hfe=input.clean_hfe(),
                remove_second_column=input.remove_second_col()
            )
            data.set(df)
            data_cleaned.set(True)
            analysis_run.set(False)
            data_info.set(results)
            ui.notification_show("Dataset cleaned successfully", type="success")
        except Exception as e:
            ui.notification_show(f"Error cleaning dataset: {str(e)}", type="error", duration=None)
    
    @reactive.Effect
    @reactive.event(input.btn_analyze)
    def _():
        if data() is None:
            ui.notification_show("Please upload a dataset first", type="warning")
            return
        
        if not data_cleaned():
            ui.notification_show("It's recommended to clean the dataset before analysis", type="info")
        
        try:
            df = data()
            results = []
            
            # Find HFE columns
            hfe_columns = []
            for col in df.columns:
                if 'hfe' in str(col).lower():
                    hfe_columns.append(col)
            
            if not hfe_columns:
                ui.notification_show("No HFE mutation columns found in the dataset", type="warning")
                return
            
            # Add information about dataset
            results.append("Dataset columns:")
            for i, col in enumerate(df.columns):
                results.append(f"  {i}: {col}")
            
            # Run selected analysis
            analysis_type = input.analysis_type()
            
            if analysis_type in ["basic", "all"]:
                basic_results = analyze_dataset(df)
                results.extend(basic_results)
            
            if analysis_type in ["hh_risk", "all"]:
                df, hh_results = analyze_hh_risk(df, hfe_columns)
                results.extend(hh_results)
            
            if analysis_type in ["diagnosis", "all"]:
                df, diag_results = analyze_diagnosis_associations(df, hfe_columns)
                results.extend(diag_results)
            
            if analysis_type in ["hardy_weinberg", "all"]:
                df, hwe_results, hwe_results_list = analyze_hardy_weinberg(df, hfe_columns)
                results.extend(hwe_results)
                plots_hardy_weinberg.set(generate_hardy_weinberg_plots(hwe_results_list))
            
            # Generate plots
            plots_genotype.set(generate_genotype_distribution_plot(df, hfe_columns))
            
            # Generate new plots for age, gender, and diagnosis-gender relationships
            age_plots, age_diagnostics = generate_genotype_by_age_plot(df, hfe_columns)
            plots_age.set(age_plots)
            diagnostics_age.set(age_diagnostics)
            
            gender_plots, gender_diagnostics = generate_genotype_by_gender_plot(df, hfe_columns)
            plots_gender.set(gender_plots)
            diagnostics_gender.set(gender_diagnostics)
            
            diag_gender_plots, diag_gender_diagnostics = generate_genotype_diagnosis_gender_plot(df, hfe_columns)
            plots_diagnosis_gender.set(diag_gender_plots)
            diagnostics_diagnosis_gender.set(diag_gender_diagnostics)
            
            if analysis_type in ["hh_risk", "all"]:
                plot_risk.set(generate_risk_distribution_plot(df))
            
            if analysis_type in ["diagnosis", "all"]:
                plot_diagnosis.set(generate_diagnosis_association_plot(df))
                plot_liver.set(generate_liver_disease_plot(df))
            
            # Update data with analysis results
            data.set(df)
            analysis_results_text.set(results)
            analysis_run.set(True)
            ui.notification_show("Analysis completed", type="success")
        
        except Exception as e:
            ui.notification_show(f"Error during analysis: {str(e)}", type="error", duration=None)
            import traceback
            error_traceback = traceback.format_exc()
            analysis_results_text.set([f"Error during analysis: {str(e)}", "", "Traceback:", error_traceback])
            analysis_run.set(True)
    
    @output
    @render.text
    def dataset_info():
        if data() is None:
            return "No dataset loaded"
        return "\n".join(data_info())
    
    @output
    @render.data_frame
    def data_preview():
        if data() is None:
            return None
        return render.DataGrid(data().head(10), width="100%")
    
    @output
    @render.text
    def analysis_results():
        if not analysis_run():
            return "Run analysis to see results"
        return "\n".join(analysis_results_text())
    
    @output
    @render.ui
    def genotype_plots():
        if not plots_genotype():
            return ui.p("No genotype plots available. Run analysis first.")
        
        plot_htmls = []
        for name, img_str in plots_genotype():
            plot_htmls.append(ui.tags.div(
                ui.tags.h4(name),
                ui.tags.img(src=f"data:image/png;base64,{img_str}", width="100%", style="max-width: 800px;"),
                ui.br(), ui.br()
            ))
        
        return ui.tags.div(*plot_htmls)
    
    @output
    @render.ui
    def genotype_age_plots():
        if not plots_age() or len(plots_age()) == 0:
            diagnostics = diagnostics_age()
            if diagnostics:
                # Show the most relevant diagnostic information
                key_messages = [msg for msg in diagnostics if any(term in msg for term in 
                              ['column', 'found', 'Sample', 'Age range', 'Error'])]
                
                # Limit to 5 most important messages
                if len(key_messages) > 5:
                    key_messages = key_messages[:5]
                    
                return ui.tags.div(
                    ui.p("No age plots available. Diagnostic information:"),
                    ui.tags.ul(*[ui.tags.li(msg) for msg in key_messages]),
                    ui.p("See Analysis Results tab for complete diagnostics.")
                )
            else:
                return ui.p("No age plots available. This could be because age data is missing or there aren't enough samples to create meaningful visualizations.")
        
        plot_htmls = []
        for name, img_str in plots_age():
            plot_htmls.append(ui.tags.div(
                ui.tags.h4(name),
                ui.tags.img(src=f"data:image/png;base64,{img_str}", width="100%", style="max-width: 800px;"),
                ui.br(), ui.br()
            ))
        
        return ui.tags.div(*plot_htmls)
    
    @output
    @render.ui
    def genotype_gender_plots():
        if not plots_gender() or len(plots_gender()) == 0:
            diagnostics = diagnostics_gender()
            if diagnostics:
                # Show the most relevant diagnostic information
                key_messages = [msg for msg in diagnostics if any(term in msg for term in 
                              ['column', 'found', 'Sample', 'distribution', 'Error'])]
                
                # Limit to 5 most important messages
                if len(key_messages) > 5:
                    key_messages = key_messages[:5]
                    
                return ui.tags.div(
                    ui.p("No gender plots available. Diagnostic information:"),
                    ui.tags.ul(*[ui.tags.li(msg) for msg in key_messages]),
                    ui.p("See Analysis Results tab for complete diagnostics.")
                )
            else:
                return ui.p("No gender plots available. This could be because gender data is missing or there aren't enough samples to create meaningful visualizations.")
        
        plot_htmls = []
        for name, img_str in plots_gender():
            plot_htmls.append(ui.tags.div(
                ui.tags.h4(name),
                ui.tags.img(src=f"data:image/png;base64,{img_str}", width="100%", style="max-width: 800px;"),
                ui.br(), ui.br()
            ))
        
        return ui.tags.div(*plot_htmls)
    
    @output
    @render.ui
    def genotype_diagnosis_gender_plots():
        if not plots_diagnosis_gender() or len(plots_diagnosis_gender()) == 0:
            diagnostics = diagnostics_diagnosis_gender()
            if diagnostics:
                # Show the most relevant diagnostic information
                key_messages = [msg for msg in diagnostics if any(term in msg for term in 
                              ['column', 'found', 'Sample', 'distribution', 'Error'])]
                
                # Limit to 5 most important messages
                if len(key_messages) > 5:
                    key_messages = key_messages[:5]
                    
                return ui.tags.div(
                    ui.p("No diagnosis-gender plots available. Diagnostic information:"),
                    ui.tags.ul(*[ui.tags.li(msg) for msg in key_messages]),
                    ui.p("See Analysis Results tab for complete diagnostics.")
                )
            else:
                return ui.p("No diagnosis-gender plots available. This could be because diagnosis or gender data is missing or there aren't enough samples to create meaningful visualizations.")
        
        plot_htmls = []
        for name, img_str in plots_diagnosis_gender():
            plot_htmls.append(ui.tags.div(
                ui.tags.h4(name),
                ui.tags.img(src=f"data:image/png;base64,{img_str}", width="100%", style="max-width: 800px;"),
                ui.br(), ui.br()
            ))
        
        return ui.tags.div(*plot_htmls)
    
    @output
    @render.ui
    def risk_plot():
        if not plot_risk():
            return ui.p("No risk distribution plot available. Run HH Risk analysis first.")
        
        return ui.tags.div(
            ui.tags.img(src=f"data:image/png;base64,{plot_risk()}", width="100%", style="max-width: 800px;")
        )
    
    @output
    @render.ui
    def diagnosis_plot():
        if not plot_diagnosis():
            return ui.p("No diagnosis association plot available. Run Diagnosis Association analysis first.")
        
        return ui.tags.div(
            ui.tags.img(src=f"data:image/png;base64,{plot_diagnosis()}", width="100%", style="max-width: 800px;")
        )
    
    @output
    @render.ui
    def liver_disease_plot():
        if not plot_liver():
            return ui.p("No liver disease plot available. Run Diagnosis Association analysis first.")
        
        return ui.tags.div(
            ui.tags.img(src=f"data:image/png;base64,{plot_liver()}", width="100%", style="max-width: 800px;")
        )
    
    @output
    @render.ui
    def hardy_weinberg_plots():
        if not plots_hardy_weinberg():
            return ui.p("No Hardy-Weinberg plots available. Run Hardy-Weinberg analysis first.")
        
        plot_htmls = []
        for mutation, img_str in plots_hardy_weinberg():
            plot_htmls.append(ui.tags.div(
                ui.tags.h4(mutation),
                ui.tags.img(src=f"data:image/png;base64,{img_str}", width="100%", style="max-width: 800px;"),
                ui.br(), ui.br()
            ))
        
        return ui.tags.div(*plot_htmls)
    
    @reactive.Effect
    @reactive.event(data_cleaned)
    def _():
        # This effect will run whenever data_cleaned changes, but we'll manage
        # the button states through the disabled attribute in the UI
        pass
        
    @reactive.Effect
    @reactive.event(data)
    def _():
        # This effect will run whenever data changes, but we'll manage
        # the button states through the disabled attribute in the UI
        pass
        
    # Helper functions to determine button states
    @reactive.calc
    def formatted_button_disabled():
        return not (data() is not None and data_cleaned())
    
    @reactive.calc
    def download_button_disabled():
        return data() is None
    
    @output
    @render.ui
    def formatted_download_button():
        return ui.download_button(
            "download_formatted", 
            "Save Formatted Dataset", 
            class_="btn-outline-secondary mt-2", 
            disabled=formatted_button_disabled()
        )
    
    @output
    @render.ui
    def data_download_button():
        return ui.download_button(
            "download_data", 
            "Download Processed Data", 
            class_="btn-info", 
            disabled=download_button_disabled()
        )
    
    @output
    @render.download(filename="processed_dataset.xlsx")
    def download_data():
        # If no data, return None
        if data() is None:
            return None
            
        # Create temporary file name
        import os
        import tempfile
        
        temp_path = os.path.join(tempfile.gettempdir(), "processed_dataset.xlsx")
        
        # Get data and format ID column
        df = data().copy()
        
        # Find ID column (likely first column)
        if 'id' in df.columns:
            id_col = 'id'
        elif any(col.lower() == 'id' for col in df.columns):
            id_col = next(col for col in df.columns if col.lower() == 'id')
        else:
            # Assuming the first column is the ID
            id_col = df.columns[0]
            
        # Fix ID formatting if needed
        if df[id_col].dtype in [np.int64, np.float64]:
            df[id_col] = df[id_col].astype('Int64')
            df[id_col] = df[id_col].astype(str)
            df[id_col] = df[id_col].replace(['nan', '<NA>'], None)
        
        # Apply zfill to preserve leading zeros
        mask = df[id_col].notna() & (df[id_col] != '')
        if any(mask):
            # Check for ID format
            sample_ids = df[id_col].dropna().head(10).tolist()
            id_lengths = [len(str(id_val)) for id_val in sample_ids if str(id_val) != 'nan' and str(id_val) != '']
            zfill_length = 10 if not id_lengths or max(id_lengths) > 9 else 9
            # Apply consistent formatting
            df.loc[mask, id_col] = df.loc[mask, id_col].str.zfill(zfill_length)
        
        # Write to Excel with formatting for ID column
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            
            # Apply text format to ID column to preserve leading zeros
            worksheet = writer.sheets['Sheet1']
            id_col_index = list(df.columns).index(id_col) + 1  # +1 because Excel is 1-indexed
            for cell in worksheet.iter_cols(min_col=id_col_index, max_col=id_col_index, min_row=2):
                for x in cell:
                    x.number_format = '@'
        
        # Return the path as a string
        return temp_path
    
    @output
    @render.download(filename="formatted_dataset.xlsx")
    def download_formatted():
        # If no data, return None
        if data() is None:
            return None
            
        # Create temporary file name
        import os
        import tempfile
        from openpyxl.styles import Font
        
        temp_path = os.path.join(tempfile.gettempdir(), "formatted_dataset.xlsx")
        
        # Get data and format ID column
        df = data().copy()
        
        # Find ID column (likely first column)
        if 'id' in df.columns:
            id_col = 'id'
        elif any(col.lower() == 'id' for col in df.columns):
            id_col = next(col for col in df.columns if col.lower() == 'id')
        else:
            # Assuming the first column is the ID
            id_col = df.columns[0]
            
        # Fix ID formatting - convert to string
        if df[id_col].dtype in [np.int64, np.float64]:
            # For numeric columns, only convert non-NaN values to string
            df[id_col] = df[id_col].astype('Int64')  # nullable integer type
            df[id_col] = df[id_col].astype(str)
            # Replace 'nan' or '<NA>' strings with None
            df[id_col] = df[id_col].replace(['nan', '<NA>'], None)
        
        # Apply zfill to non-null values to preserve leading zeros
        mask = df[id_col].notna() & (df[id_col] != '')
        if any(mask):
            # Determine ID format based on sample
            sample_ids = df[id_col].dropna().head(10).tolist()
            id_lengths = [len(str(id_val)) for id_val in sample_ids if str(id_val) != 'nan' and str(id_val) != '']
            zfill_length = 10 if not id_lengths or max(id_lengths) > 9 else 9
            # Apply zfill for consistent length
            df.loc[mask, id_col] = df.loc[mask, id_col].str.zfill(zfill_length)
        
        # Write to Excel with formatting
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            
            # Apply formatting to Excel worksheet
            worksheet = writer.sheets['Sheet1']
            
            # Create a bold font object directly
            bold_font = Font(bold=True)
            
            # Format headers in bold
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = bold_font
            
            # Set ID column to text format (to preserve leading zeros)
            id_col_index = list(df.columns).index(id_col) + 1  # +1 because Excel is 1-indexed
            for cell in worksheet.iter_cols(min_col=id_col_index, max_col=id_col_index, min_row=2):
                for x in cell:
                    x.number_format = '@'
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Return the path as a string
        return temp_path

# Create the Shiny app
app = App(app_ui, server)

# Run the app
if __name__ == "__main__":
    print("Starting app on http://127.0.0.1:8095")
    # Use only supported parameters
    app.run(host="127.0.0.1", port=8095) 